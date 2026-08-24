#!/usr/bin/env python3
"""
Import a programme spreadsheet into the job ledger.

    python scripts/import_job_ledger.py Texas_Invoice_Tracker_deposits.xlsx \
        --client "KBP Foods" --program "Project Red"

Each worksheet becomes a category, taken from the tab name ("Parking Lots" →
parking, "Roof" → roof), and every row is graded by what it actually carries.
Run with --dry-run first: it prints the grades without writing anything, which
is the only way to see that a sheet of 28 sites holds 9 invoiced jobs before
the numbers are in the database.

The spreadsheet itself is not committed. It carries the amounts billed to a
client, which are confidential to both sides of that relationship.
"""

import argparse
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.services import job_ledger  # noqa: E402


def category_for(sheet_name: str) -> str | None:
    """
    The kind of work a tab holds, or None when the tab is not about kind.

    Two workbooks, two conventions: the Texas tracker names its tabs for the
    work ("Parking Lots", "Roof"), the KFC tracker names them for the state
    ("GA", "TX", "NJ", "MI", "NY"). Filing a record under category "ga" is not
    wrong so much as useless, and it throws away the state, which is the field
    a regional page filters on.
    """
    if job_ledger.state_from_sheet_name(sheet_name):
        return None
    name = sheet_name.strip().lower()
    if "roof" in name:
        return "roof"
    if "park" in name or "lot" in name:
        return "parking"
    return name.replace(" ", "_")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--client", default=None)
    parser.add_argument("--program", default=None)
    parser.add_argument("--tenant", default="default")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("openpyxl is needed to read .xlsx: pip install openpyxl", file=sys.stderr)
        return 2

    if not args.workbook.exists():
        print(f"No such file: {args.workbook}", file=sys.stderr)
        return 2

    workbook = openpyxl.load_workbook(args.workbook, data_only=True)
    everything: list[dict] = []

    for sheet_name in workbook.sheetnames:
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        records = job_ledger.read_program_sheet(
            rows,
            client=args.client,
            program=args.program,
            category=category_for(sheet_name),
            state=job_ledger.state_from_sheet_name(sheet_name),
            source_document=args.workbook.name,
        )
        if not records:
            print(f"  {sheet_name}: no header row with a store number and address — skipped")
            continue

        summary = job_ledger.summarise(records)
        label = (
            f"state '{job_ledger.state_from_sheet_name(sheet_name)}'"
            if job_ledger.state_from_sheet_name(sheet_name)
            else f"category '{category_for(sheet_name)}'"
        )
        print(f"\n  {sheet_name} → {label}")
        for grade_name in reversed(job_ledger.EVIDENCE_ORDER):
            bucket = summary["by_evidence"].get(grade_name)
            if not bucket:
                continue
            money = f"  ${bucket['invoiced_dollars']}" if bucket["invoiced_cents"] else ""
            mark = "publishable" if bucket["publishable"] else "NOT publishable"
            print(f"    {grade_name:<10} {bucket['count']:>4}  ({mark}){money}")
        everything.extend(records)

    total = job_ledger.summarise(everything)
    print(
        f"\n  {total['records']} records, {total['publishable']} of them backed by an invoice."
    )

    if args.dry_run:
        print("  --dry-run: nothing written.")
        return 0

    from app.database import SessionLocal  # noqa: PLC0415
    from app.models import ClientJobRecord  # noqa: PLC0415

    session = SessionLocal()
    created = upgraded = 0
    try:
        for record in everything:
            existing = (
                session.query(ClientJobRecord)
                .filter(
                    ClientJobRecord.tenant_id == args.tenant,
                    ClientJobRecord.store_number == record["store_number"],
                    ClientJobRecord.category == record["category"],
                    ClientJobRecord.state == record["state"],
                )
                .first()
            )
            if existing is None:
                session.add(ClientJobRecord(tenant_id=args.tenant, **record))
                created += 1
                continue
            # A re-import may only strengthen a record, never demote one.
            if job_ledger.rank(record["evidence"]) > job_ledger.rank(existing.evidence):
                existing.evidence = record["evidence"]
                upgraded += 1
            for field, value in record.items():
                if field == "evidence" or value in (None, ""):
                    continue
                if getattr(existing, field, None) in (None, ""):
                    setattr(existing, field, value)
        session.commit()
    finally:
        session.close()

    print(f"  Written: {created} new, {upgraded} upgraded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

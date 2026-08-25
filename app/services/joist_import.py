"""
joist_import.py — read the Joist workbook, and grade every row before it counts.

WHAT THIS IS FOR
────────────────
Joist is the system the business invoiced through after Kickserv and before
InvoiceFly. It was cancelled in early 2026. 52 documents were recovered from
the mailbox — dates, clients, amounts, sender notes and document links — and
this reads the workbook they were assembled into.

It matters because every other record in this repository stops on 2022-04-04.
`recentWork.js` currently covers 2022-2026 with five entries scraped out of
correspondence. This is the real derivation that replaces them.

THE ONE RULE THAT WILL COST MONEY IF IT IS WRONG
────────────────────────────────────────────────
The workbook has a sheet of "other invoices/bills — RV park, gravel, hauling,
equipment". Those are things the business BOUGHT. They are accounts payable.
A naive import that sees a sheet called invoices and adds up the amount column
produces a revenue figure inflated by every load of gravel the company ever
paid for.

So the sheet a row came from decides its DIRECTION as well as its grade, and
`PAYABLE` rows are excluded from every earned total this module reports. That
is checked by a test rather than left to whoever reads the code next.

THE SECOND RULE: A BID IS NOT A JOB
───────────────────────────────────
The "Paving Jobs & Leads" sheet is inquiries, bid invitations and prices sent —
Shakerag WRF, Greenville County speed humps, Hull Street Self Storage, the
Chester shopping center. Impressive names, and almost none of them evidence
work performed. They grade `requested` or `quoted`, neither of which is
publishable, for the same reason the KBP tracker's 246 "Not Started" rows are
not a portfolio.

WHAT CANNOT BE RECOVERED
────────────────────────
Line-item scope. Joist's document pages are app-only, so the workbook carries
totals and the sender's covering note but not what the money actually bought.
That means a row can say a job was invoiced at a figure; it cannot say the
figure was for paving rather than for materials or a change order. Any page
built from this may state the client and the amount, and must not describe the
work in detail it does not have.

CHESTER
───────
Chester, South Carolina is 350 miles from Chester, Virginia, and the business
is headquartered in the second one. A city-name match merges them and files a
South Carolina shopping centre as home-market work. `AMBIGUOUS_CITIES` names
the traps; a row in one of those towns must carry a state or it resolves to
nothing.
"""

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Optional

logger = logging.getLogger(__name__)

# ── Direction: did money come in, or go out? ─────────────────────────────────
RECEIVABLE = "receivable"
PAYABLE = "payable"

# ── Evidence, sharing the ladder the rest of the system uses ─────────────────
REQUESTED = "requested"
QUOTED = "quoted"
CONTRACTED = "contracted"
INVOICED = "invoiced"

PUBLISHABLE = frozenset({CONTRACTED, INVOICED})

#: Sheet name -> (direction, default grade). Matched case-insensitively on a
#: substring, because a workbook's tab names drift with every rebuild.
SHEET_RULES = (
    ("estimate", (RECEIVABLE, QUOTED)),
    ("invoice", (RECEIVABLE, INVOICED)),
    ("receipt", (RECEIVABLE, INVOICED)),
    ("lead", (RECEIVABLE, REQUESTED)),
    ("job", (RECEIVABLE, REQUESTED)),
    ("bid", (RECEIVABLE, REQUESTED)),
    # Must be tested BEFORE the generic 'invoice' rule when a tab is called
    # "Other invoices/bills": the word 'other' is what makes it payable.
    ("other", (PAYABLE, INVOICED)),
    ("bill", (PAYABLE, INVOICED)),
    ("expense", (PAYABLE, INVOICED)),
)

#: Checked first, so "Other invoices/bills" cannot match 'invoice' and be read
#: as money earned.
PAYABLE_FIRST = ("other", "bill", "expense", "payable", "supplier", "vendor")

#: Words in a status or note that lift a quote to a contract.
_ACCEPTED = re.compile(r"\b(accepted|approved|signed|awarded|won|countersigned)\b", re.I)
#: Words that mean the opposite. A lost bid is `requested`, never higher.
_LOST = re.compile(r"\b(lost|declined|rejected|no bid|not awarded|cancell?ed)\b", re.I)

#: Towns that exist in more than one state this company works in. A row in one
#: of these MUST carry a state or its location is not resolved.
AMBIGUOUS_CITIES = {
    "chester": ("VA", "SC"),
    "lancaster": ("SC", "PA"),
    "greenville": ("SC", "NC", "TX"),
    "richmond": ("VA", "TX"),
    "columbia": ("SC", "MO"),
    "florence": ("SC", "AL"),
    # Beaufort SC (Lowcountry, where this company works) and Beaufort NC (the
    # Crystal Coast) are both real and 350 miles apart. Caught because the
    # lead list holds a "Holiday Inn Beaufort reseal" with no state on it.
    "beaufort": ("SC", "NC"),
    "jackson": ("GA", "MS", "MI"),
    "franklin": ("VA", "TN"),
    "aurora": ("IL", "CO"),
    "kansas city": ("MO", "KS"),
}


def direction_and_grade(sheet_name: Any) -> tuple[str, str]:
    """
    What a sheet's rows mean. Payable patterns are tested first on purpose.

    "Other invoices/bills" contains the word `invoice`. Matching that first
    would classify every gravel delivery and equipment payment as revenue.
    """
    name = str(sheet_name or "").strip().lower()
    for token in PAYABLE_FIRST:
        if token in name:
            return (PAYABLE, INVOICED)
    for token, result in SHEET_RULES:
        if token in name:
            return result
    logger.warning("sheet %r matched no rule — treated as an unpublishable lead", sheet_name)
    return (RECEIVABLE, REQUESTED)


def grade_row(sheet_name: Any, status: Any = None, note: Any = None) -> str:
    """
    The sheet's default grade, adjusted by anything the row itself says.

    A status column beats a sheet name in both directions: an estimate marked
    accepted is `contracted`, and an estimate marked lost is `requested` rather
    than `quoted`, because a price nobody took is not a price the market paid.
    """
    _, grade = direction_and_grade(sheet_name)
    text = f"{status or ''} {note or ''}"
    if _LOST.search(text):
        return REQUESTED
    if grade == QUOTED and _ACCEPTED.search(text):
        return CONTRACTED
    return grade


def to_cents(value: Any) -> Optional[int]:
    """
    Money as whole cents through Decimal. Never float.

    Handles the shapes a spreadsheet produces: 1234.5, '$1,234.50', '(500)'
    for a credit, and '' for a cell nobody filled in.
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        text = str(value)
    else:
        text = str(value).strip()
        negative = text.startswith("(") and text.endswith(")")
        text = re.sub(r"[^0-9.\-]", "", text)
        if negative:
            text = f"-{text}"
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return int((Decimal(text) * 100).to_integral_value())
    except (InvalidOperation, ValueError):
        logger.warning("could not read %r as money", value)
        return None


def to_date(value: Any) -> Optional[str]:
    """An ISO date, or None. Never today's date as a stand-in."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text[:20], fmt).date().isoformat()
        except ValueError:
            continue
    # A range or a composite: "2024-05-13 to 06-12", "2024-01-29; 2024-10-10",
    # "2025-05-22 invite; 2025-07-16 status". Nearly every row on the leads and
    # bills sheets is one of these, because a lead is a correspondence trail
    # rather than a document with one date on it.
    #
    # Take the EARLIEST date present, which is when the thing started. Taking
    # the last would date a job by the most recent chase email and quietly move
    # 2023 work into 2024.
    found = re.findall(r"\b(\d{4})-(\d{2})-(\d{2})\b", text)
    if found:
        return min("-".join(f) for f in found)
    logger.warning("could not read %r as a date", value)
    return None


def resolve_location(city: Any, state: Any = None) -> dict[str, Optional[str]]:
    """
    A town, and whether it is safe to act on without a state.

    Chester SC is 350 miles from Chester VA and the company is headquartered in
    the second. A row in an ambiguous town with no state resolves to `None` and
    is flagged, rather than silently becoming home-market work.
    """
    town = re.sub(r"\s+", " ", str(city or "").strip())
    code = str(state or "").strip().upper()[:2] or None
    if not town:
        return {"city": None, "state": code, "ambiguous": False}
    candidates = AMBIGUOUS_CITIES.get(town.lower())
    if candidates and not code:
        return {"city": town, "state": None, "ambiguous": True, "could_be": list(candidates)}
    return {"city": town, "state": code, "ambiguous": False}


def read_rows(rows: Iterable[dict], sheet_name: str) -> list[dict]:
    """
    One sheet's rows to graded records. Pure — takes dicts, returns dicts.

    Deliberately does not open a file: the reader can then be tested against a
    handful of literal rows rather than against a workbook that has to be
    regenerated every time the format changes.
    """
    direction, _ = direction_and_grade(sheet_name)
    out: list[dict] = []
    for row in rows:
        lower = {str(k or "").strip().lower(): v for k, v in row.items()}

        def pick(*names):
            for n in names:
                if lower.get(n) not in (None, ""):
                    return lower[n]
            return None

        # Column aliases, per sheet-author rather than per specification. A
        # workbook a person built names the same thing four ways across four
        # tabs — client / contact / from, date / date sent / date(s), amount /
        # total / "quoted $ (from email)". Matching only the tidy names read
        # the entire leads sheet as 19 empty rows: no client, no reference, no
        # amount, and no error, because every lookup legitimately missed.
        # A silent empty column is the failure mode to design against here.
        def like(*fragments):
            """First column whose NAME contains one of these fragments."""
            for fragment in fragments:
                for key, value in lower.items():
                    if fragment in key and value not in (None, ""):
                        return value
            return None

        amount_cents = to_cents(like("amount", "total", "quoted", "value", "price"))
        status = like("status", "paid", "delivery note", "notes")
        note = like("scope", "message", "what it is", "note")
        record = {
            "sheet": sheet_name,
            "direction": direction,
            "evidence": grade_row(sheet_name, status, note),
            "ref": (str(like("estimate #", "invoice #", "number", "job / lead", "ref", "type") or "").strip() or None),
            "client": (str(like("client", "contact", "customer", "from", "name") or "").strip() or None),
            "date": to_date(like("date")),
            "amount_cents": amount_cents,
            "note": (str(note or "").strip() or None),
            "status": (str(status or "").strip() or None),
            "brand": (str(like("business name") or "").strip() or None),
            "link": (str(like("full document", "link", "url") or "").strip() or None),
            **resolve_location(like("city", "town"), like("state", " st")),
        }
        out.append(record)
    return out


def distinct_jobs(records: list[dict]) -> list[dict]:
    """
    One record per JOB, where several documents describe the same one.

    Joist issues an invoice and then a payment receipt for the same work, and
    every one of those documents carries the job's FULL total. Summing the
    documents therefore counts a job once per receipt it generated.

    In the Carolina archive that inflates the figure by $51,750 on a $155,300
    book — a third again — through three jobs alone:

        daven89        9,000  x3   (6557, 6563, 6565)
        Ascent Realty 13,500  x3   (6571, 6572, 6580)
        John           6,750  x2   (6583, 6585)

    texasProgram.js states the same rule for the same reason: an advance and a
    final are never summed, because Greenville reads 17,949 twice and is ONE
    job. This is that rule, applied structurally rather than by hand.

    The key is (client, amount). Two genuinely separate jobs for one client at
    an identical price would be collapsed wrongly — but that is far rarer than
    an invoice followed by its receipt, and understating is the safer error.
    """
    seen: dict[tuple, dict] = {}
    for record in records:
        if record.get("amount_cents") is None:
            # No amount, nothing to double count. Kept as its own row.
            seen[(id(record),)] = record
            continue
        key = (str(record.get("client") or "").strip().lower(), record["amount_cents"])
        first = seen.get(key)
        if first is None:
            seen[key] = {**record, "documents": [record.get("ref")]}
        else:
            first.setdefault("documents", [first.get("ref")]).append(record.get("ref"))
    return list(seen.values())


def summarise(records: list[dict]) -> dict[str, Any]:
    """
    What the workbook establishes. Earned totals EXCLUDE payable rows.

    The `payable_cents` figure is reported separately and never added in. A
    single number combining what was billed out with what was paid for gravel
    means nothing and looks authoritative, which is the worst combination.
    """
    receivable = [r for r in records if r["direction"] == RECEIVABLE]
    payable = [r for r in records if r["direction"] == PAYABLE]
    # Deduplicated FIRST. Summing raw documents counts a job once per receipt.
    invoiced = distinct_jobs([r for r in receivable if r["evidence"] == INVOICED])
    invoiced_documents = [r for r in receivable if r["evidence"] == INVOICED]

    dates = sorted(r["date"] for r in records if r["date"])
    return {
        "documents": len(records),
        "receivable": len(receivable),
        "payable": len(payable),
        "by_grade": {
            g: sum(1 for r in receivable if r["evidence"] == g)
            for g in (REQUESTED, QUOTED, CONTRACTED, INVOICED)
        },
        # The only earned figure: invoiced receivables, ONE PER JOB.
        "invoiced_cents": sum(r["amount_cents"] or 0 for r in invoiced),
        "invoiced_jobs": len(invoiced),
        "invoiced_documents": len(invoiced_documents),
        # Surfaced so the gap between documents and jobs is visible rather than
        # quietly absorbed — it is the difference between $207,050 and $155,300.
        "duplicate_documents": len(invoiced_documents) - len(invoiced),
        # Reported so it is visible, never folded into the line above.
        "payable_cents": sum(r["amount_cents"] or 0 for r in payable),
        "publishable": sum(1 for r in receivable if r["evidence"] in PUBLISHABLE),
        "ambiguous_locations": [
            {"ref": r["ref"], "city": r["city"], "could_be": r.get("could_be")}
            for r in records if r.get("ambiguous")
        ],
        "first_date": dates[0] if dates else None,
        "last_date": dates[-1] if dates else None,
    }


#: A row is the header when it carries several of these. Assuming row 1 is the
#: header is wrong for any workbook a human built: this one opens with a title,
#: a provenance line and a blank, and the columns start on row 4. Reading row 1
#: as the header silently produced a table whose every column was named None.
_HEADER_WORDS = frozenset({
    "client", "customer", "contact", "date", "date sent", "date(s)", "amount",
    "total", "number", "invoice #", "estimate #", "type", "status", "notes",
    "job / lead", "from", "what it is",
})


def find_header(rows: list[tuple]) -> int:
    """Index of the header row. Falls back to the first populated row."""
    for index, row in enumerate(rows[:20]):
        labels = {str(c).strip().lower() for c in row if c not in (None, "")}
        if len(labels & _HEADER_WORDS) >= 3:
            return index
    for index, row in enumerate(rows):
        if sum(1 for c in row if c not in (None, "")) >= 3:
            return index
    return 0


def _is_total_row(row: dict) -> bool:
    """
    A spreadsheet's own footer, which is not a document.

    "Total (where amount known):" sits in the date column of every sheet here.
    Counting it as a record inflates the document count and, worse, gives that
    phantom row a grade.
    """
    joined = " ".join(str(v) for v in row.values() if v not in (None, "")).lower()
    return joined.startswith("total") or "total (where amount known)" in joined


def read_workbook(path: str) -> dict[str, Any]:
    """The whole workbook. Every sheet, graded, plus the summary."""
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    for sheet in book.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        head = find_header(rows)
        header = [str(c or "").strip() for c in rows[head]]
        dicts = [
            dict(zip(header, r))
            for r in rows[head + 1:]
            if any(c not in (None, "") for c in r)
        ]
        dicts = [d for d in dicts if not _is_total_row(d)]
        records.extend(read_rows(dicts, sheet.title))
    return {"records": records, "summary": summarise(records)}
